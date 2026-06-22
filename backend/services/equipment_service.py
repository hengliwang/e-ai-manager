from sqlalchemy.orm import Session
from models.equipment import Equipment, EquipmentPhoto, AuditLog, FieldConfig
from models.inspection_task import InspectionTask, TaskPhoto, TaskDefect
from models.defect_order import DefectOrder
from schemas.equipment import EquipmentCreate, EquipmentUpdate, FieldConfigCreate, FieldConfigUpdate
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def get_equipment_list(db: Session, skip: int = 0, limit: int = 20,
                       category: str = None, equipment_type: str = None,
                       keyword: str = None, customer_name: str = None):
    query = db.query(Equipment)
    if category:
        query = query.filter(Equipment.category == category)
    if equipment_type:
        query = query.filter(Equipment.equipment_type == equipment_type)
    if customer_name:
        query = query.filter(Equipment.customer_name == customer_name)
    if keyword:
        query = query.filter(
            (Equipment.equipment_name.contains(keyword)) |
            (Equipment.asset_code.contains(keyword)) |
            (Equipment.address_detail.contains(keyword))
        )
    total = query.count()
    items = query.order_by(Equipment.updated_at.desc()).offset(skip).limit(limit).all()
    return total, items


def get_equipment(db: Session, equipment_id: int):
    return db.query(Equipment).filter(Equipment.id == equipment_id).first()


def create_equipment(db: Session, data: EquipmentCreate, user_id: int):
    equipment = Equipment(**data.model_dump(), created_by=user_id)
    db.add(equipment)
    db.commit()
    db.refresh(equipment)
    return equipment


def update_equipment(db: Session, equipment_id: int, data: EquipmentUpdate, user_id: int):
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        return None
    for key, new_val in data.model_dump(exclude_unset=True).items():
        old_val = getattr(equipment, key)
        if str(old_val) != str(new_val):
            audit = AuditLog(
                equipment_id=equipment_id,
                field_name=key,
                old_value=str(old_val) if old_val else None,
                new_value=str(new_val),
                modified_by=user_id
            )
            db.add(audit)
        setattr(equipment, key, new_val)
    db.commit()
    db.refresh(equipment)
    return equipment


def delete_equipment(db: Session, equipment_id: int):
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        return None

    # 级联删除：找到该设备的所有巡检任务
    task_ids = db.query(InspectionTask.id).filter(InspectionTask.equipment_id == equipment_id).all()
    task_id_list = [t[0] for t in task_ids]

    if task_id_list:
        # 删除消缺工单（关联到任务缺陷）
        db.query(DefectOrder).filter(
            DefectOrder.task_defect_id.in_(
                db.query(TaskDefect.id).filter(TaskDefect.task_id.in_(task_id_list))
            )
        ).delete(synchronize_session=False)

        # 删除任务缺陷
        db.query(TaskDefect).filter(TaskDefect.task_id.in_(task_id_list)).delete(synchronize_session=False)

        # 删除任务照片
        db.query(TaskPhoto).filter(TaskPhoto.task_id.in_(task_id_list)).delete(synchronize_session=False)

        # 删除巡检任务
        db.query(InspectionTask).filter(InspectionTask.equipment_id == equipment_id).delete(synchronize_session=False)

    # 删除直接关联到该设备的消缺工单
    db.query(DefectOrder).filter(DefectOrder.equipment_id == equipment_id).delete(synchronize_session=False)

    # 删除审计日志
    db.query(AuditLog).filter(AuditLog.equipment_id == equipment_id).delete(synchronize_session=False)

    # 删除设备照片
    db.query(EquipmentPhoto).filter(EquipmentPhoto.equipment_id == equipment_id).delete(synchronize_session=False)

    # 删除设备
    db.delete(equipment)
    db.commit()
    return equipment


# ===== 字段配置管理 =====

def get_field_configs(db: Session, include_disabled: bool = False):
    query = db.query(FieldConfig)
    if not include_disabled:
        query = query.filter(FieldConfig.is_active == "active")
    return query.order_by(FieldConfig.sort_order).all()


def get_field_config(db: Session, config_id: int):
    return db.query(FieldConfig).filter(FieldConfig.id == config_id).first()


def create_field_config(db: Session, data: FieldConfigCreate):
    config = FieldConfig(**data.model_dump())
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def update_field_config(db: Session, config_id: int, data: FieldConfigUpdate):
    config = db.query(FieldConfig).filter(FieldConfig.id == config_id).first()
    if not config:
        return None
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(config, key, value)
    db.commit()
    db.refresh(config)
    return config


def delete_field_config(db: Session, config_id: int):
    config = db.query(FieldConfig).filter(FieldConfig.id == config_id).first()
    if config:
        db.delete(config)
        db.commit()
    return config


def manage_field_options(db: Session, config_id: int, action: str, option: dict = None, old_value: str = None):
    """管理字段配置的选项列表: add, update, remove, toggle"""
    config = db.query(FieldConfig).filter(FieldConfig.id == config_id).first()
    if not config or config.options is None:
        return None

    options = config.options

    if action == "add":
        if option:
            options.append(option)
    elif action == "update":
        if old_value and option:
            for i, opt in enumerate(options):
                opt_val = opt["value"] if isinstance(opt, dict) else opt
                if opt_val == old_value:
                    options[i] = option
                    break
    elif action == "remove":
        if old_value:
            options = [o for o in options if (o["value"] if isinstance(o, dict) else o) != old_value]
    elif action == "toggle":
        if old_value:
            for opt in options:
                if isinstance(opt, dict) and opt.get("value") == old_value:
                    opt["active"] = not opt.get("active", True)
                    break

    config.options = options
    db.commit()
    db.refresh(config)
    return config


# ===== 导入导出 =====

EQUIPMENT_CORE_FIELDS = [
    "category", "equipment_type", "asset_code", "equipment_name",
    "cabinet_model", "factory_number", "line_name", "station_name",
    "operation_date", "manufacturer", "province", "city", "district",
    "street", "address_detail", "longitude", "latitude",
    "customer_name", "remark",
]


def export_equipment_excel(db: Session) -> io.BytesIO:
    """导出全部设备到 Excel"""
    configs = get_field_configs(db)
    equipment_list = db.query(Equipment).order_by(Equipment.updated_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "设备档案"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # 构建列: 所有激活的字段配置
    active_configs = [c for c in configs if c.is_active == "active"]
    headers = [c.field_label for c in active_configs]
    field_names = [c.field_name for c in active_configs]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    for row_idx, equip in enumerate(equipment_list, 2):
        for col_idx, field_name in enumerate(field_names, 1):
            val = getattr(equip, field_name, None)
            if val is None and equip.extra_fields:
                val = equip.extra_fields.get(field_name, "")
            if val is None:
                val = ""
            elif isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")

    # 自动列宽
    for col_idx in range(1, len(headers) + 1):
        max_width = len(str(headers[col_idx - 1])) * 2 + 4
        for row_idx in range(2, min(len(equipment_list) + 2, 100)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                # 中文字符算2个宽度
                width = sum(2 if ord(c) > 127 else 1 for c in str(cell_val))
                max_width = max(max_width, width)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_equipment_excel(db: Session, file_bytes: bytes, user_id: int) -> dict:
    """从 Excel 批量导入设备"""
    wb = Workbook()
    # openpyxl 从 bytes 加载
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # 读取表头行
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())

    # 建立 label → field_name 映射
    configs = get_field_configs(db, include_disabled=True)
    label_to_field = {c.field_label: c.field_name for c in configs}

    success = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue

        data = {}
        extra_fields = {}
        for col_idx, header in enumerate(headers):
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            val_str = str(val).strip()
            if not val_str:
                continue

            field_name = label_to_field.get(header, header)
            if field_name in EQUIPMENT_CORE_FIELDS:
                data[field_name] = val_str
            else:
                extra_fields[field_name] = val_str

        if not data.get("equipment_name") or not data.get("category") or not data.get("equipment_type"):
            errors.append(f"第{row_idx}行: 缺少必填字段(设备名称/设备大类/设备类型)")
            continue

        if extra_fields:
            data["extra_fields"] = extra_fields

        try:
            create_equipment(db, EquipmentCreate(**data), user_id)
            success += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    return {"success": success, "total": success + len(errors), "errors": errors}
